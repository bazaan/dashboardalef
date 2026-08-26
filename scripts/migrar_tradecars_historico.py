# -*- coding: utf-8 -*-
"""
Migra el histórico de Trade Cars del Excel del asesor a `tradecars_funnel_leads`.

    python scripts/migrar_tradecars_historico.py            # dry-run: no escribe
    python scripts/migrar_tradecars_historico.py --escribir  # sube a Supabase

Es idempotente: cada fila lleva un `import_key` estable (hoja + nº de fila +
teléfono), así que volver a correrlo actualiza en vez de duplicar.

No calcula etapa ni fecha del funnel: eso lo hacen las columnas GENERATED de la
BD. Tampoco resuelve zona ni prioridad de marca: lo hace el trigger
`tradecars_funnel_autocompletar`. Aquí sólo se traduce el Excel.

Requiere: SUPABASE_URL y SUPABASE_KEY en el .env del proyecto.
"""
import sys, os, io, re, json, hashlib, datetime, unicodedata, urllib.request
from collections import Counter

# La consola de Windows viene en cp1252 y este script imprime acentos.
try:
    sys.stdout.reconfigure(encoding='utf-8')
except (AttributeError, ValueError):
    pass

try:
    import openpyxl
except ImportError:
    sys.exit('Falta openpyxl:  pip install openpyxl')

RAIZ  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL = os.environ.get('TC_EXCEL', r'C:\Users\elrob\Desktop\BASE COMPRAS - MIGUEL C.xlsx')
LOTE  = 400

STATUS_VALIDOS = {
    'NO CONTACTADO', 'NO INTERESADO', 'EN SEGUIMIENTO',
    'CITA', 'CITA ASISTIDA', 'CONCRETADA',
}

# Motivos tal como quedaron sembrados en tradecars_funnel_motivos.
MOTIVOS = {
    'PRECIO': 'Precio',
    'NO RECIBIMOS EL MODELO': 'No recibimos el modelo',
    'YA LO VENDIO': 'Ya lo vendio',
    'SU DEUDA ES MAYOR AL PRECIO OFERTADO': 'Su deuda es mayor al precio ofertado',
    'NO RESPONDE': 'No responde',
}

# El Excel no tiene columna de canal: el CRM sí. Se deduce sólo cuando la
# campaña lo dice sin ambigüedad; en el resto se deja vacío antes que inventarlo.
CANAL_POR_CAMPANA = [
    (r'\bTIK ?TOK\b',                    'TikTok'),
    (r'\bWTP\b|\bWHATSAPP\b|\bWSP\b',    'WhatsApp'),
    (r'\bIG\b|\bINSTAGRAM\b',            'Instagram'),
    (r'\bFB\b|\bFACEBOOK\b|MARKETPLACE', 'Facebook'),
    (r'\bWEB\b',                         'Web'),
]


# ══════════════════ utilidades ══════════════════

def norm(v):
    s = str(v if v is not None else '').strip()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return ' '.join(s.upper().split())


def texto(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ('', '-', '.', '#N/A', '#VALUE!', 'null', 'None'):
        return None
    return s


def fecha(v):
    if v is None or v == '':
        return None
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    s = str(v).strip()
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
    if m:
        return '%s-%s-%s' % m.groups()
    m = re.match(r'^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$', s)
    if m:
        return '%s-%02d-%02d' % (m.group(3), int(m.group(2)), int(m.group(1)))
    # Serial de Excel
    if re.match(r'^\d{5}$', s):
        try:
            base = datetime.date(1899, 12, 30)
            return (base + datetime.timedelta(days=int(s))).isoformat()
        except ValueError:
            return None
    return None


def numero(v, entero=False):
    if v is None or v == '':
        return None
    s = str(v).strip().replace('S/', '').replace('$', '').replace(',', '').replace(' ', '')
    if not re.match(r'^-?\d+(\.\d+)?$', s):
        return None
    try:
        return int(float(s)) if entero else round(float(s), 2)
    except ValueError:
        return None


def telefono(v):
    d = ''.join(ch for ch in str(v if v is not None else '') if ch.isdigit())
    if len(d) > 9 and d.startswith('51'):
        d = d[2:]
    return d or None


def env(clave):
    ruta = os.path.join(RAIZ, '.env')
    if not os.path.exists(ruta):
        return None
    for linea in io.open(ruta, encoding='utf-8'):
        m = re.match(r'^%s=(.*)$' % re.escape(clave), linea.strip())
        if m:
            return m.group(1).strip().strip('"').strip("'")
    return None


# ══════════════════ lectura del Excel ══════════════════

# Nombre de la columna en el Excel -> columna en la BD.
MAPA = {
    'FECHA DE DERIVACION':      ('fecha_derivacion', fecha),
    'FECHA DE LLEGADA':         ('fecha_llegada', fecha),
    'CAMPANA':                  ('campana', texto),
    '¿DEUDA?':                  ('tiene_deuda', texto),
    'DEUDA?':                   ('tiene_deuda', texto),
    'BANCO':                    ('banco', texto),
    'NOMBRE':                   ('contacto_nombre', texto),
    'NUMERO':                   ('contacto_telefono', telefono),
    'CORREO':                   ('correo', texto),
    'DISTRITO':                 ('distrito', texto),
    'ZONAS':                    ('zona', texto),
    'ASESORES':                 ('asesor', texto),
    'PLACA':                    ('placa', texto),
    'MARCA':                    ('marca', texto),
    'MODELO':                   ('modelo', texto),
    'VERSION':                  ('version', texto),
    'ANO':                      ('anio', texto),
    'KM':                       ('kilometraje', lambda v: numero(v, True)),
    'PERFIL COINCIDE':          ('perfil_coincide', texto),
    'STATUS':                   ('status', texto),
    'MONTO PROPUESTA INICIAL':  ('monto_propuesta_inicial', numero),
    'MONTO MEJORADO':           ('monto_mejorado', numero),
    'EXPECTATIVA CLIENTE':      ('expectativa_cliente', numero),
    'FECHA DE CITA':            ('fecha_cita', fecha),
    'FECHA ULTIMO CONTACTO':    ('fecha_ultimo_contacto', fecha),
    'FECHA DE COMPRA':          ('fecha_compra', fecha),
    '# DE CONTACTOS':           ('num_contactos', lambda v: numero(v, True)),
    'FEEDBACK':                 ('feedback', texto),
    '¿POR QUE NO SE CONCRETO?': ('motivo_no_cita', texto),
    'POR QUE NO SE CONCRETO?':  ('motivo_no_cita', texto),
    'MARCA // PRIORIDAD':       ('marca_prioridad', lambda v: numero(v, True)),
}

# Todas las claves que puede llevar una fila. PostgREST exige que todos los
# objetos de un insert masivo tengan exactamente las mismas claves.
COLUMNAS = sorted({v[0] for v in MAPA.values()} | {
    'canal_origen', 'lead_origen_tabla', 'import_key', 'observaciones',
})


def leer(hoja, etiqueta):
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    if hoja not in wb.sheetnames:
        hoja = [s for s in wb.sheetnames if norm(s) == norm(hoja)][0]
    ws = wb[hoja]
    it = ws.iter_rows(values_only=True)
    cabecera = next(it)
    cols = {}
    for i, h in enumerate(cabecera):
        k = norm(h)
        if k in MAPA:
            cols[i] = MAPA[k]

    faltan = {v[0] for v in MAPA.values()} - {v[0] for v in cols.values()}
    filas, saltadas = [], Counter()

    for n, r in enumerate(it, start=2):
        fila = {c: None for c in COLUMNAS}
        for i, (destino, conv) in cols.items():
            if i < len(r):
                fila[destino] = conv(r[i])

        # Una fila sin contacto ni fechas no aporta nada al funnel.
        if not (fila['contacto_telefono'] or fila['contacto_nombre']):
            saltadas['sin contacto'] += 1
            continue
        if not (fila['fecha_derivacion'] or fila['fecha_cita'] or fila['fecha_compra']):
            saltadas['sin ninguna fecha'] += 1
            continue

        # Status fuera de la lista cerrada: se guarda igual (el dashboard lo
        # marca en rojo), pero se cuenta para el informe.
        s = norm(fila['status'])
        if s and s not in STATUS_VALIDOS:
            saltadas['status no reconocido (se guarda igual)'] += 1
        fila['status'] = s or None

        p = norm(fila['perfil_coincide'])
        fila['perfil_coincide'] = p if p in ('SI', 'NO') else None

        fila['motivo_no_cita'] = MOTIVOS.get(norm(fila['motivo_no_cita']), fila['motivo_no_cita'])

        d = norm(fila['tiene_deuda'])
        fila['tiene_deuda'] = 'SI' if d in ('SI', 'S', 'YES', 'TRUE', '1') else ('NO' if d else None)

        if fila['marca_prioridad'] not in (1, 2, 3):
            fila['marca_prioridad'] = None

        camp = norm(fila['campana'])
        for patron, canal in CANAL_POR_CAMPANA:
            if re.search(patron, camp):
                fila['canal_origen'] = canal
                break

        fila['lead_origen_tabla'] = etiqueta
        fila['import_key'] = 'xl:%s:%s' % (
            etiqueta,
            hashlib.sha1(('%d|%s|%s' % (n, fila['contacto_telefono'] or '',
                                        fila['fecha_derivacion'] or '')).encode()).hexdigest()[:16],
        )
        filas.append(fila)

    return filas, saltadas, faltan


# ══════════════════ subida ══════════════════

def preflight(base, key):
    """Comprueba que sql/tradecars_funnel.sql ya se corrio.

    Sin la columna `import_key` la migracion no seria idempotente, y sin los
    catalogos la zona y la prioridad quedarian vacias en las ~8.700 filas.
    """
    cab = {'apikey': key, 'Authorization': 'Bearer ' + key}
    faltan = []
    for tabla, campo in [('tradecars_funnel_leads', 'import_key'),
                         ('tradecars_zonificacion', 'clave'),
                         ('tradecars_marcas', 'clave')]:
        url = '%s/rest/v1/%s?select=%s&limit=1' % (base.rstrip('/'), tabla, campo)
        try:
            urllib.request.urlopen(urllib.request.Request(url, headers=cab), timeout=30).read()
        except urllib.error.HTTPError:
            faltan.append('%s.%s' % (tabla, campo))
    return faltan


def subir(filas, base, key):
    url = base.rstrip('/') + '/rest/v1/tradecars_funnel_leads?on_conflict=import_key'
    cab = {
        'apikey': key,
        'Authorization': 'Bearer ' + key,
        'Content-Type': 'application/json',
        'Prefer': 'resolution=merge-duplicates,return=minimal',
    }
    ok = err = 0
    for i in range(0, len(filas), LOTE):
        lote = filas[i:i + LOTE]
        req = urllib.request.Request(
            url, data=json.dumps(lote, ensure_ascii=False).encode('utf-8'),
            headers=cab, method='POST')
        try:
            with urllib.request.urlopen(req, timeout=180) as r:
                r.read()
            ok += len(lote)
            print('   lote %-4d  %d/%d' % (i // LOTE + 1, ok, len(filas)))
        except urllib.error.HTTPError as e:
            err += len(lote)
            print('   ERROR lote %d (%s): %s' % (i // LOTE + 1, e.code, e.read().decode()[:400]))
    return ok, err


def main():
    escribir = '--escribir' in sys.argv
    print('Excel:', EXCEL)
    print('Modo :', 'ESCRITURA' if escribir else 'dry-run (no escribe nada)')

    total, todas = [], []
    for hoja, etiqueta in [('BASE LEADS', 'EXCEL BASE LEADS'), ('HISTORICO', 'EXCEL HISTORICO')]:
        filas, saltadas, faltan = leer(hoja, etiqueta)
        print('\n=== %s' % hoja)
        print('   a migrar : %d' % len(filas))
        for k, v in saltadas.most_common():
            print('   %-38s %d' % (k, v))
        if faltan:
            print('   columnas del mapa que esa hoja no tiene: %s' % ', '.join(sorted(faltan)))
        todas.extend(filas)

    # Dos filas del Excel podrían generar la misma clave; se queda la última.
    porclave = {}
    for f in todas:
        porclave[f['import_key']] = f
    total = list(porclave.values())
    if len(total) != len(todas):
        print('\n   %d filas con import_key repetida: se conserva la última' % (len(todas) - len(total)))

    print('\n──────────────────────────────────────────')
    print('TOTAL a migrar: %d filas' % len(total))
    meses = Counter(f['fecha_derivacion'][:7] for f in total if f['fecha_derivacion'])
    print('rango: %s .. %s (%d meses)' % (min(meses), max(meses), len(meses)))
    con = lambda c: sum(1 for f in total if f[c])
    for c in ['perfil_coincide', 'status', 'distrito', 'marca', 'campana',
              'canal_origen', 'marca_prioridad', 'fecha_compra', 'motivo_no_cita']:
        print('   %-18s %5d  (%.0f%%)' % (c, con(c), 100.0 * con(c) / max(len(total), 1)))

    if not escribir:
        print('\nDry-run. Para subirlo:  python %s --escribir' % os.path.basename(__file__))
        return

    base, key = env('SUPABASE_URL'), env('SUPABASE_KEY')
    if not base or not key:
        sys.exit('Faltan SUPABASE_URL / SUPABASE_KEY en el .env')
    faltan = preflight(base, key)
    if faltan:
        sys.exit('Falta correr sql/tradecars_funnel.sql en Supabase. '
                 'No se encontro: ' + ', '.join(faltan))

    print('\nSubiendo a %s ...' % base)
    ok, err = subir(total, base, key)
    print('\nListo: %d subidas, %d con error' % (ok, err))


if __name__ == '__main__':
    main()
